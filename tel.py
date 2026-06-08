import sqlite3
import io
import csv
import traceback
import inspect
import os
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from dotenv import load_dotenv
from nicegui import app, ui

# ==========================================
# 0. Environment & Logging Setup
# ==========================================

load_dotenv()

PORT = int(os.getenv('PORT', 8080))
ROOT_PATH = os.getenv('ROOT_PATH', '')
LOG_LEVEL_STR = os.getenv('LOG_LEVEL', 'INFO').upper()

# راه‌اندازی سیستم لاگ‌نویسی (5 مگابایت، 20 فایل بکاپ)
log_level = getattr(logging, LOG_LEVEL_STR, logging.INFO)
logger = logging.getLogger('ExirpooyanPhonebook')
logger.setLevel(log_level)

log_handler = RotatingFileHandler('phonebook.log', maxBytes=5 * 1024 * 1024, backupCount=20, encoding='utf-8')
log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
log_handler.setFormatter(log_formatter)
if not logger.handlers:
    logger.addHandler(log_handler)

logger.info("Starting Exirpooyan Phonebook Application...")

# ==========================================
# 1. Database Setup & Manager
# ==========================================

DB_FILE = 'phonebook.db'

def init_db():
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        c.execute('''CREATE TABLE IF NOT EXISTS contacts
                     (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                      location TEXT, unit TEXT, name TEXT, phone1 TEXT, phone2 TEXT)''')
        
        c.execute("PRAGMA table_info(contacts)")
        columns = [info[1] for info in c.fetchall()]
        if 'location' not in columns:
            c.execute("ALTER TABLE contacts ADD COLUMN location TEXT DEFAULT 'دفتر مرکزی'")
            logger.info("Database altered: added 'location' column.")
            
        c.execute('''CREATE TABLE IF NOT EXISTS settings
                     (key TEXT PRIMARY KEY, value TEXT)''')
        
        c.execute("SELECT value FROM settings WHERE key='username'")
        if not c.fetchone():
            c.execute("INSERT INTO settings (key, value) VALUES ('username', 'admin')")
            c.execute("INSERT INTO settings (key, value) VALUES ('password', 'admin')")
            logger.info("Default admin credentials created.")
        
        conn.commit()
    except Exception as e:
        logger.critical(f"Database Init Error: {e}", exc_info=True)
    finally:
        if 'conn' in locals(): 
            conn.close()

init_db()

def get_db_connection():
    conn = sqlite3.connect(DB_FILE, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def get_credentials():
    try:
        conn = get_db_connection()
        user = conn.execute("SELECT value FROM settings WHERE key='username'").fetchone()['value']
        pwd = conn.execute("SELECT value FROM settings WHERE key='password'").fetchone()['value']
        return user, pwd
    except Exception as e:
        logger.error(f"Error fetching credentials: {e}", exc_info=True)
        return "admin", "admin"
    finally:
        if 'conn' in locals(): 
            conn.close()

def update_credentials(new_user, new_pwd):
    try:
        conn = get_db_connection()
        conn.execute("UPDATE settings SET value=? WHERE key='username'", (new_user,))
        conn.execute("UPDATE settings SET value=? WHERE key='password'", (new_pwd,))
        conn.commit()
        logger.info("Admin credentials updated successfully.")
    except Exception as e:
        logger.error(f"Error updating credentials: {e}", exc_info=True)
        raise
    finally:
        if 'conn' in locals(): 
            conn.close()

def get_all_contacts():
    try:
        conn = get_db_connection()
        contacts = conn.execute("SELECT * FROM contacts ORDER BY location, unit, name").fetchall()
        return [dict(c) for c in contacts]
    except Exception as e:
        logger.error(f"Error fetching contacts: {e}", exc_info=True)
        return []
    finally:
        if 'conn' in locals(): 
            conn.close()

# ==========================================
# 2. Global CSS, Theme & UI Helpers
# ==========================================

def apply_global_styles():
    ui.add_head_html('''  
<style>
    @import url('https://cdn.jsdelivr.net/gh/rastikerdar/vazirmatn@v33.0.0/Vazirmatn-font-face.css') layer(default);
    body { 
        font-family: 'Vazirmatn', 'Calibri', 'B Nazanin', 'Tahoma', sans-serif !important; 
        direction: rtl; 
        margin: 0; 
        padding: 0; 
        transition: background-color 0.3s ease; 
    } 
    .q-table__container { 
        direction: rtl; 
        transition: all 0.3s ease; 
    } 
    .q-table th { 
        text-align: right !important; 
        font-size: 15px !important; 
        font-weight: 800 !important; 
    } 
    .q-table td { 
        text-align: right !important; 
        font-size: 14px !important; 
        font-weight: 500; 
    } 
    .majlesi-card { 
        border-radius: 12px; 
        box-shadow: 0 8px 24px -4px rgba(0,0,0,0.1); 
        transition: all 0.3s ease; 
    }

    /* Light Mode Styles */
    body.body--light { background-color: #f0f2f5; }
    body.body--light .majlesi-card { background-color: #ffffff; }
    body.body--light .q-table th { background-color: #f8fafc; color: #1e293b; }
    body.body--light .unit-header { background-color: #e2e8f0; color: #1e293b; border-right: 4px solid #3b82f6; }
    
    /* Dark Mode Styles */
    body.body--dark { background-color: #121212; }
    body.body--dark .majlesi-card { background-color: #1e1e1e; border: 1px solid #333; }
    body.body--dark .q-table th { background-color: #2d2d2d; color: #e2e8f0; }
    body.body--dark .q-table td { color: #cbd5e1; }
    body.body--dark .unit-header { background-color: #333333; color: #e2e8f0; border-right: 4px solid #60a5fa; }
    
    .header-bg { background: linear-gradient(90deg, #0f172a 0%, #1e3a8a 100%); color: white; border-bottom: 3px solid #3b82f6; }
    .unit-header { padding: 8px 16px; border-radius: 6px; font-weight: bold; margin-top: 16px; margin-bottom: 8px; }
</style>
''', shared=True)

def theme_toggle_button():
    dark = ui.dark_mode()
    
    def set_theme(value):
        dark.set_value(value)
        # ذخیره ترجیح تم کاربر در کوکی مرورگر
        app.storage.browser['dark_mode'] = value

    with ui.button(icon='palette').props('flat round color=white').classes('ml-2'): 
        with ui.menu().classes('min-w-[120px]'): 
            ui.menu_item('روشن (Light)', on_click=lambda: set_theme(False)).classes('font-bold') 
            ui.menu_item('تاریک (Dark)', on_click=lambda: set_theme(True)).classes('font-bold') 
            ui.menu_item('سیستم (Auto)', on_click=lambda: set_theme(None)).classes('font-bold')

def load_user_theme():
    """بارگذاری ترجیح تم کاربر از کوکی مرورگر به محض ورود به صفحه"""
    try:
        dark = ui.dark_mode()
        saved_theme = app.storage.browser.get('dark_mode', None)
        if saved_theme is not None:
            dark.set_value(saved_theme)
    except Exception as e:
        logger.error(f"Error loading saved theme: {e}")

# ==========================================
# 3. Native Excel/CSV Handlers
# ==========================================

def generate_sample_excel(): 
    wb = openpyxl.Workbook() 
    ws = wb.active 
    ws.title = "Exirpooyan_Format" 
    ws.append(['محل کار', 'واحد/بخش', 'نام و نام خانوادگی', 'تلفن اول', 'تلفن دوم']) 
    ws.append(['دفتر تهران', 'فناوری اطلاعات', 'علی احمدی', '09120000000', '02188888888']) 
    ws.append(['کارخانه اراک', 'تولید', 'رضا کریمی', '08633333333', 'داخلی 112']) 
    output = io.BytesIO() 
    wb.save(output) 
    return output.getvalue()

def export_database(): 
    wb = openpyxl.Workbook() 
    ws = wb.active 
    ws.title = "Exirpooyan_Contacts" 
    ws.append(['محل کار', 'واحد/بخش', 'نام و نام خانوادگی', 'تلفن اول', 'تلفن دوم']) 
    for c in get_all_contacts(): 
        ws.append([c['location'], c['unit'], c['name'], c['phone1'], c['phone2']]) 
    output = io.BytesIO() 
    wb.save(output) 
    logger.info("Database exported to Excel.") 
    return output.getvalue()

async def handle_upload(e): 
    try: 
        logger.info(f"Starting file upload process: {getattr(e, 'name', 'unknown')}") 
        file_obj = getattr(e, 'content', None) or getattr(e, 'file', None) 
        if file_obj is None: 
            file_obj = e

        read_result = file_obj.read()
        file_bytes = await read_result if inspect.isawaitable(read_result) else read_result
        filename = getattr(e, 'name', 'unknown.xlsx').lower()

        rows = []
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
        elif filename.endswith('.csv'):
            text = file_bytes.decode('utf-8-sig')
            rows = list(csv.reader(io.StringIO(text)))
        else:
            ui.notify('❌ فرمت فایل پشتیبانی نمی‌شود. فقط csv یا xlsx', type='negative', position='top')
            logger.warning(f"Unsupported file format uploaded: {filename}")
            return

        if not rows or len(rows) < 2:
            ui.notify('❌ فایل خالی است یا فقط ردیف عنوان دارد.', type='warning', position='top')
            return

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        data_rows = rows[1:]

        def find_idx(keywords, default):
            for i, h in enumerate(headers):
                if any(k in h for k in keywords): 
                    return i
            return default if default < len(headers) else -1

        idx_loc  = find_idx(['محل', 'شهر', 'دفتر', 'کارخانه'], 0)
        idx_unit = find_idx(['واحد', 'بخش', 'دپارتمان'], 1)
        idx_name = find_idx(['نام', 'شخص', 'مخاطب'], 2)
        idx_p1   = find_idx(['تلفن 1', 'تلفن اول', 'اصلی', 'موبایل'], 3)
        idx_p2   = find_idx(['تلفن 2', 'تلفن دوم', 'داخلی', 'ثابت'], 4)

        success_count, failed_count = 0, 0
        conn = get_db_connection()
        
        for r in data_rows:
            try:
                def get_val(idx):
                    if idx != -1 and idx < len(r) and r[idx] is not None:
                        val = str(r[idx]).strip()
                        return val if val.lower() not in ['none', 'nan', 'null'] else ''
                    return ''

                loc, unit, name, p1, p2 = get_val(idx_loc), get_val(idx_unit), get_val(idx_name), get_val(idx_p1), get_val(idx_p2)
                loc = loc or 'نامشخص'
                unit = unit or 'نامشخص'

                if name:
                    conn.execute("INSERT INTO contacts (location, unit, name, phone1, phone2) VALUES (?, ?, ?, ?, ?)",
                                 (loc, unit, name, p1, p2))
                    success_count += 1
                else:
                    if any([loc != 'نامشخص', unit != 'نامشخص', p1, p2]): 
                        failed_count += 1
            except Exception as row_e:
                logger.debug(f"Row parsing failed: {r} | Error: {row_e}")
                failed_count += 1 
                
        conn.commit()
        conn.close()
        
        if success_count > 0:
            msg = f'✅ {success_count} مخاطب با موفقیت افزوده شد.'
            if failed_count > 0: 
                msg += f' ({failed_count} ردیف نامعتبر نادیده گرفته شد)'
            ui.notify(msg, type='positive', position='top', timeout=5000)
            logger.info(f"Upload successful: {success_count} added, {failed_count} failed.")
            ui.navigate.reload()
        else:
            ui.notify('❌ هیچ دیتای معتبری برای ذخیره پیدا نشد!', type='negative', position='top')
            logger.warning("Upload processed but no valid data found.")
            
    except Exception as ex:
        logger.error("Upload Error Detail", exc_info=True)
        ui.notify(f'❌ خطای سیستمی در پردازش فایل.', type='negative', position='top')

# ==========================================
# 4. Pure Python Levenshtein & Normalizer Search Engine
# ==========================================

def normalize_persian(text):
    """استانداردسازی حروف عربی/فارسی و فاصله‌ها جهت یکسان‌سازی جستجو در رم"""
    if not text:
        return ""
    text = str(text).strip().lower()
    replacements = {
        'ي': 'ی',
        'ك': 'ک',
        'ة': 'ه',
        'أ': 'ا',
        'إ': 'ا',
        'آ': 'ا',
        '‌': ' ',  # حذف نیم‌فاصله
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text

def levenshtein_distance(s1, s2):
    """محاسبه دقیق فاصله لوناشتاین بین دو رشته"""
    if len(s1) < len(s2):
        return levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    
    previous_row = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        current_row = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = previous_row[j + 1] + 1
            deletions = current_row[j] + 1
            substitutions = previous_row[j] + (c1 != c2)
            current_row.append(min(insertions, deletions, substitutions))
        previous_row = current_row
        
    return previous_row[-1]

def levenshtein_similarity(s1, s2):
    """تبدیل فاصله لوناشتاین به درصد تشابه بین 0.0 تا 1.0"""
    if not s1 or not s2:
        return 0.0
    dist = levenshtein_distance(s1, s2)
    max_len = max(len(s1), len(s2))
    return 1.0 - (dist / max_len)

def search_contacts_in_memory(query, all_contacts):
    """
    موتور جستجوی پیشرفته بر روی داده‌های بارگذاری شده در رم.
    پشتیبانی از:
    1. تطبیق جزیی (بخش، تلفن، نام و محل کار) در هر کجای متن.
    2. تطبیق توکن‌ها (پشتیبانی از جابه‌جایی اسامی مثل "کریمی رضا").
    3. محاسبه شباهت فازی لوناشتاین در صورت عدم وجود نتیجه مستقیم.
    """
    query_norm = normalize_persian(query)
    if not query_norm:
        return [], [], False

    direct_matches = []
    fuzzy_candidates = []
    query_tokens = [t for t in query_norm.split() if t]

    for c in all_contacts:
        name_norm = normalize_persian(c.get('name', ''))
        phone1_norm = normalize_persian(c.get('phone1', ''))
        phone2_norm = normalize_persian(c.get('phone2', ''))
        unit_norm = normalize_persian(c.get('unit', ''))
        loc_norm = normalize_persian(c.get('location', ''))

        # ۱. تطبیق زیررشته‌ای مستقیم در تمام فیلدها (شروع، وسط، پایان)
        if (query_norm in name_norm or 
            query_norm in phone1_norm or 
            query_norm in phone2_norm or 
            query_norm in unit_norm or 
            query_norm in loc_norm):
            direct_matches.append(c)
            continue

        # ۲. تطبیق توکن‌ها (پشتیبانی از جابه‌جایی کلمات ورودی مثل "کریمی رضا")
        if query_tokens:
            all_tokens_in_name = all(token in name_norm for token in query_tokens)
            all_tokens_in_unit = all(token in unit_norm for token in query_tokens)
            if all_tokens_in_name or all_tokens_in_unit:
                direct_matches.append(c)
                continue

        # ۳. بخش فازی (لوناشتاین) - فقط برای فیلد نام
        sim_score = levenshtein_similarity(query_norm, name_norm)
        
        token_scores = []
        name_tokens = [t for t in name_norm.split() if t]
        for q_tok in query_tokens:
            for n_tok in name_tokens:
                token_scores.append(levenshtein_similarity(q_tok, n_tok))
        
        max_token_score = max(token_scores) if token_scores else 0.0
        best_score = max(sim_score, max_token_score)

        # آستانه پذیرش تطابق فازی (بیش از ۵۰ درصد تشابه)
        if best_score >= 0.5:
            fuzzy_candidates.append((best_score, c))

    if direct_matches:
        return direct_matches, [], False

    # مرتب‌سازی نتایج فازی بر اساس بیشترین شباهت و نمایش ۵ مورد اول
    fuzzy_candidates.sort(key=lambda x: x[0], reverse=True)
    top_fuzzy = [item[1] for item in fuzzy_candidates[:5]]
    
    return [], top_fuzzy, len(top_fuzzy) > 0

# ==========================================
# 5. UI Pages
# ==========================================

@ui.page('/') 
def user_page(): 
    apply_global_styles()
    load_user_theme()

    with ui.header().classes('header-bg p-4 flex justify-between items-center shadow-lg'):
        with ui.row().classes('items-center gap-3'):
            ui.icon('contact_phone', size='36px', color='blue-3')
            ui.label('دفترچه تلفن شرکت اکسیرپویان').classes('text-2xl font-bold tracking-wide')
        
        with ui.row().classes('items-center gap-2'):
            theme_toggle_button()
            ui.button('دریافت فایل اکسل', icon='download', color='green-5', 
                      on_click=lambda: ui.download(export_database(), 'Exirpooyan_Phonebook.xlsx')).props('rounded').classes('font-bold shadow-md')
            ui.button('مدیریت', icon='admin_panel_settings', color='white', 
                      on_click=lambda: ui.navigate.to('/login')).props('rounded outline').classes('text-white font-bold')

    with ui.column().classes('w-full max-w-5xl mx-auto p-4 mt-6 gap-4'):
        
        # فیلد سرچ (راست‌چین شده با تراز مناسب فیلدها و نشانگر تایپ فارسی)
        search_input = ui.input(
            label='جستجو بر اساس نام، شماره تلفن، بخش یا محل کار...',
            placeholder='مثال: علی، فناوری اطلاعات، ۰۹۱۲، دفتر تهران و ...'
        ).classes('w-full').props('outlined clearable icon=search').style('direction: rtl; text-align: right;')

        # تابع کمکی برای پاک کردن کامل سرچ و بازگشت به لیست
        def clear_search():
            search_input.value = ''
            render_contacts.refresh()

        # کانتینر نمایش نتایج (واکنش‌گرا و متصل به رفرش امن)
        @ui.refreshable
        def render_contacts():
            # بارگذاری سریع کل اطلاعات در رم
            all_contacts = get_all_contacts()
            
            # دریافت مقدار فیلد سرچ به طور مستقیم
            query = (search_input.value or '').strip()
            
            if not query:
                # ۱. حالت عادی: نمایش کلیه مخاطبین به صورت دسته‌بندی شده
                ui.label('جهت مشاهده شماره تماس‌ها، روی "محل کار" مورد نظر کلیک کنید:').classes('text-xl mb-2 font-bold text-gray-700 dark:text-gray-200')
                if not all_contacts:
                    ui.label('هیچ اطلاعاتی در سیستم ثبت نشده است.').classes('text-gray-500 italic text-center w-full mt-10 text-lg')
                    return
                    
                grouped_data = {}
                for c in all_contacts:
                    loc, unit = c['location'] or 'نامشخص', c['unit'] or 'نامشخص'
                    if loc not in grouped_data: 
                        grouped_data[loc] = {}
                    if unit not in grouped_data[loc]: 
                        grouped_data[loc][unit] = []
                    grouped_data[loc][unit].append(c)
                    
                for loc, units in grouped_data.items():
                    with ui.expansion(loc, icon='business').classes('w-full majlesi-card text-2xl font-bold mb-3 overflow-hidden border-t-4 border-blue-500'):
                        for unit, unit_contacts in units.items():
                            ui.label(unit).classes('unit-header text-lg w-full block')
                            columns = [
                                {'name': 'name', 'label': 'نام و نام خانوادگی', 'field': 'name', 'align': 'right'},
                                {'name': 'phone1', 'label': 'تلفن اصلی', 'field': 'phone1', 'align': 'left'},
                                {'name': 'phone2', 'label': 'تلفن داخلی/دوم', 'field': 'phone2', 'align': 'left'},
                            ]
                            ui.table(columns=columns, rows=unit_contacts, row_key='id').classes('w-full shadow-none mb-4').props('flat bordered dense')
            
            else:
                # ۲. حالت فیلتر یا جستجوی هوشمند فازی در حافظه رم
                direct_matches, fuzzy_matches, is_fuzzy = search_contacts_in_memory(query, all_contacts)
                
                # هدر و دکمه صریح بازگشت به لیست اصلی در بالای نتایج
                with ui.row().classes('w-full justify-between items-center mb-4 p-3 bg-blue-50 dark:bg-slate-800 rounded-lg border border-blue-100 dark:border-slate-700'):
                    with ui.row().classes('items-center gap-2'):
                        ui.icon('search', size='24px', color='blue-6')
                        if direct_matches:
                            ui.label(f'نتایج جستجو برای "{query}" ({len(direct_matches)} مورد)').classes('text-lg font-bold text-blue-800 dark:text-blue-300')
                        else:
                            ui.label(f'جستجوی فازی برای "{query}"').classes('text-lg font-bold text-amber-800 dark:text-amber-300')
                    
                    ui.button('پاک کردن و بازگشت به لیست اصلی', icon='arrow_forward', color='red-5', on_click=clear_search)\
                        .props('flat dense').classes('font-bold')

                # نمایش نتایج مستقیم
                if direct_matches:
                    columns = [
                        {'name': 'location', 'label': 'محل کار', 'field': 'location', 'align': 'right'},
                        {'name': 'unit', 'label': 'واحد / بخش', 'field': 'unit', 'align': 'right'},
                        {'name': 'name', 'label': 'نام و نام خانوادگی', 'field': 'name', 'align': 'right'},
                        {'name': 'phone1', 'label': 'تلفن اول', 'field': 'phone1', 'align': 'left'},
                        {'name': 'phone2', 'label': 'تلفن دوم', 'field': 'phone2', 'align': 'left'},
                    ]
                    ui.table(columns=columns, rows=direct_matches, row_key='id').classes('w-full majlesi-card').props('flat bordered dense')
                
                # نمایش پیشنهادهای نزدیک (لوناشتاین) در صورت عدم وجود تطابق مستقیم
                elif is_fuzzy and fuzzy_matches:
                    ui.label('نتیجه دقیقی یافت نشد؛ نزدیک‌ترین موارد پیشنهادی به عبارت شما:').classes('text-md mb-2 text-amber-700 dark:text-amber-400')
                    columns = [
                        {'name': 'location', 'label': 'محل کار', 'field': 'location', 'align': 'right'},
                        {'name': 'unit', 'label': 'واحد / بخش', 'field': 'unit', 'align': 'right'},
                        {'name': 'name', 'label': 'نام و نام خانوادگی', 'field': 'name', 'align': 'right'},
                        {'name': 'phone1', 'label': 'تلفن اول', 'field': 'phone1', 'align': 'left'},
                        {'name': 'phone2', 'label': 'تلفن دوم', 'field': 'phone2', 'align': 'left'},
                    ]
                    ui.table(columns=columns, rows=fuzzy_matches, row_key='id').classes('w-full majlesi-card').props('flat bordered dense')
                
                # عدم وجود هیچ شباهتی در کل دیتابیس
                else:
                    with ui.column().classes('w-full items-center justify-center p-8'):
                        ui.icon('search_off', size='64px', color='grey-5')
                        ui.label(f'هیچ موردی معادل یا مشابه با "{query}" یافت نشد.').classes('text-gray-500 text-lg mt-2')

        # رندر اولیه لیست مخاطبین
        render_contacts()

        # اتصال رویداد ورودی فیلد سرچ به رفرش کانتینر
        search_input.on('update:model-value', lambda: render_contacts.refresh())


@ui.page('/login') 
def login_page(): 
    apply_global_styles()
    load_user_theme()

    def try_login():
        try:
            valid_user, valid_pwd = get_credentials()
            if username.value == valid_user and password.value == valid_pwd:
                app.storage.user['authenticated'] = True
                logger.info("Admin logged in successfully.")
                ui.navigate.to('/admin')
            else:
                logger.warning(f"Failed login attempt with username: {username.value}")
                ui.notify('❌ نام کاربری یا رمز عبور اشتباه است', type='negative', position='top')
        except Exception as e:
            logger.error("Login Error", exc_info=True)
            ui.notify('خطا در ارتباط با دیتابیس', type='negative')

    with ui.column().classes('w-full h-screen items-center justify-center'):
        with ui.page_sticky('top-right').classes('p-4'): 
            theme_toggle_button()
        
        with ui.card().classes('w-96 p-8 majlesi-card items-center gap-4'):
            ui.icon('shield_person', size='56px', color='blue-8')
            ui.label('ورود به پنل مدیریت اکسیرپویان').classes('text-2xl font-bold text-center')
            
            username = ui.input('نام کاربری').classes('w-full').props('outlined')
            password = ui.input('رمز عبور', password=True, password_toggle_button=True).classes('w-full').props('outlined')
            
            ui.button('ورود امن', on_click=try_login, color='blue-8').classes('w-full mt-4 font-bold text-lg').props('rounded size=lg')
            ui.button('بازگشت به دفترچه', on_click=lambda: ui.navigate.to('/'), color='gray-8').classes('w-full').props('flat')


@ui.page('/admin') 
def admin_page(): 
    apply_global_styles()
    load_user_theme()

    if not app.storage.user.get('authenticated', False):
        ui.navigate.to('/login')
        return

    def logout():
        app.storage.user['authenticated'] = False
        logger.info("Admin logged out.")
        ui.navigate.to('/')
        
    def clear_database():
        try:
            conn = get_db_connection()
            conn.execute("DELETE FROM contacts")
            conn.commit()
            conn.close()
            logger.warning("Database cleared by admin.")
            ui.notify('✅ تمام مخاطبین پاک شدند.', type='positive', position='top')
            ui.navigate.reload()
        except Exception as e:
            logger.error("Clear Database Error", exc_info=True)
            ui.notify('❌ خطا در پاکسازی دیتابیس', type='negative')

    with ui.dialog() as confirm_dialog, ui.card().classes('p-6 items-center majlesi-card'):
        ui.icon('warning', size='48px', color='red')
        ui.label('هشدار خطرناک').classes('text-xl font-bold text-red-600 mb-2')
        ui.label('آیا از حذف کل اطلاعات دفترچه تلفن مطمئن هستید؟').classes('text-center mb-6')
        with ui.row().classes('w-full justify-center gap-4'):
            ui.button('بله، پاک کن', color='red', on_click=clear_database).props('rounded')
            ui.button('انصراف', color='gray', on_click=confirm_dialog.close).props('rounded outline')

    with ui.header().classes('bg-slate-800 p-4 flex justify-between items-center shadow-md border-b-4 border-slate-600'):
        with ui.row().classes('items-center gap-3'):
            ui.icon('admin_panel_settings', size='32px', color='amber-4')
            ui.label('پنل مدیریت اکسیرپویان').classes('text-2xl font-bold text-white')
        
        with ui.row().classes('items-center gap-2'):
            theme_toggle_button()
            ui.button('خروج امن', icon='logout', color='red-5', on_click=logout).props('rounded')

    with ui.column().classes('w-full max-w-6xl mx-auto p-6 mt-4 gap-6'):
        with ui.row().classes('w-full gap-6'):
            
            with ui.card().classes('flex-1 majlesi-card p-6 border-t-4 border-green-500'):
                ui.label('اضافه کردن مخاطبین (آپلود اکسل/CSV)').classes('text-xl font-bold mb-2')
                ui.label('ردیف‌های نامعتبر نادیده گرفته می‌شوند.').classes('text-sm mb-6 opacity-70')
                
                ui.upload(on_upload=handle_upload, label='فایل خود را اینجا بکشید یا کلیک کنید', auto_upload=True).classes('w-full mb-4').props('accept=".csv, .xlsx, .xls"')
                
                with ui.row().classes('w-full mt-4 justify-between items-center'):
                    ui.button('دانلود فایل نمونه (خام)', icon='grid_on', color='green-7', 
                              on_click=lambda: ui.download(generate_sample_excel(), 'Exirpooyan_Sample.xlsx')).props('outline rounded')
                    ui.button('حذف کل دیتابیس', icon='delete_sweep', color='red-6', 
                              on_click=confirm_dialog.open).props('rounded')

            with ui.card().classes('w-1/3 majlesi-card p-6 border-t-4 border-blue-500'):
                ui.label('تنظیمات امنیتی').classes('text-xl font-bold mb-6')
                
                new_usr = ui.input('نام کاربری جدید', value='admin').classes('w-full mb-3').props('outlined dense')
                new_pwd = ui.input('رمز عبور جدید', password=True).classes('w-full mb-6').props('outlined dense')
                
                def change_creds():
                    try:
                        if new_usr.value.strip() and new_pwd.value.strip():
                            update_credentials(new_usr.value, new_pwd.value)
                            ui.notify('✅ یوزر و پسورد با موفقیت تغییر کرد', type='positive', position='top')
                        else:
                            ui.notify('❌ فیلدها نمی‌توانند خالی باشند', type='warning', position='top')
                    except Exception as e:
                        ui.notify('❌ خطا در ذخیره تنظیمات', type='negative')
                        
                ui.button('ذخیره رمز عبور', icon='save', color='blue-6', on_click=change_creds).classes('w-full font-bold').props('rounded')

        with ui.card().classes('w-full majlesi-card p-6 border-t-4 border-gray-400 mt-4'):
            with ui.row().classes('w-full justify-between items-center mb-4'):
                ui.label('نمایش جدولی کل اطلاعات').classes('text-xl font-bold')
                ui.button('دانلود بکاپ اکسل', icon='cloud_download', color='blue-grey-7', 
                          on_click=lambda: ui.download(export_database(), 'Exirpooyan_Backup.xlsx')).props('rounded')
            
            contacts = get_all_contacts()
            columns = [
                {'name': 'location', 'label': 'محل کار', 'field': 'location', 'align': 'right', 'sortable': True},
                {'name': 'unit', 'label': 'واحد / بخش', 'field': 'unit', 'align': 'right', 'sortable': True},
                {'name': 'name', 'label': 'نام و نام خانوادگی', 'field': 'name', 'align': 'right', 'sortable': True},
                {'name': 'phone1', 'label': 'تلفن اول', 'field': 'phone1', 'align': 'left'},
                {'name': 'phone2', 'label': 'تلفن دوم', 'field': 'phone2', 'align': 'left'},
            ]
            ui.table(columns=columns, rows=contacts, row_key='id').classes('w-full').props('flat bordered striped')

# ==========================================
# Run App
# ==========================================

# استفاده از kwargs برای ارسال تنظیمات Nginx (root_path) به uvicorn
uvicorn_kwargs = {'root_path': ROOT_PATH} if ROOT_PATH else {}

ui.run(
    title='دفترچه تلفن شرکت اکسیرپویان',
    port=PORT,
    storage_secret='exirpooyan_super_secret_key_2024',
    dark=None,  # تم اولیه از سیستم‌عامل کاربر یا کوکی ذخیره شده استخراج می‌شود
    favicon='📞',
    **uvicorn_kwargs
)
